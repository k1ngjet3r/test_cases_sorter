from openpyxl import load_workbook
from func.matcher import *
import json

# the index of the precondition
pre_index = 5

def json_directory(json_name):
    with open('json\\' + json_name) as f:
        return json.load(f)

data_sheet = json_directory('sheet_related.json')
keywords = json_directory('keywords.json')
auto_case_list = json_directory('auto_case_id.json')


def Automation_cases():
    auto_file = load_workbook('automation_cases.xlsx').active
    return [tcid[0] for tcid in auto_file.iter_rows(max_col=1, values_only=True)]


def bench_only(cell_data):
    press_button = keywords['push_button']
    cluster = keywords['cluster']
    speed_limit = keywords['speed_limit']
    expection = keywords['expection']
    for cell in cell_data[pre_index+1:pre_index+5]:
        if (matcher_slice(press_button, cell) or matcher_slice(cluster, cell) or matcher_slice(speed_limit, cell)) and not matcher_slice(expection, cell):
            return True
        return False

def ac_only(cell_data):
    ac = keywords['ac']
    ac_split = keywords['ac_split']
    for cell in cell_data[pre_index:pre_index+3]:
        if matcher_slice(ac, cell) or matcher_split(ac_split, cell):
            return True
        return False

def nav_case(cell_data):
    # Finding the navigation-related cases using TCID
    # Formatting the TCID
    tcid = [i.lower() for i in cell_data[0].split('_')]
    if 'maps' in tcid:
        return True
    return False

def call_SMS(cell_data):
    callsms = keywords['call_sms']
    if matcher_slice(callsms, cell_data[pre_index+1]):
        return True
    return False

def fuel_sim(cell_data):
    fuel = keywords['fuel_sim']
    if matcher_slice(fuel, cell_data[pre_index+1]):
        return True
    return False

def did_case(cell_data):
    did = keywords['did']
    user = keywords['user_build']
    # search DID-related case ID in test obnjective
    if matcher_slice(did, cell_data[pre_index+3]) and not matcher_slice(user, cell_data[pre_index+3]):
        return True
    return False

def user_build_only(cell_data):
    user = keywords['user_build']
    if matcher_slice(user, cell_data[pre_index+3]) or matcher_slice(user, cell_data[pre_index]):
        return True
    return False

def screen_size_13(cell_data):
    thirdteen_inch = keywords['13_inch']
    for i in range(4):
        if matcher_slice(thirdteen_inch, cell_data[pre_index+i]):
            return True
    return False

def trailer_case(cell_data):
    trailer_kw = keywords['trailer']
    for i in range(4):
        if matcher_slice(trailer_kw, cell_data[pre_index+i]):
            return True
    return False

def usb_update(cell_data):
    usb_update_kw = keywords['usb_update']
    for i in range(4):
        if matcher_slice(usb_update_kw, cell_data[pre_index+i]):
            return True
    return False