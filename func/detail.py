from openpyxl import load_workbook
from func.matcher import *
import json

# the index of the precondition
pre_index = 5

def json_directory(json_name):
    with open('json\\' + json_name) as f:
        return json.load(f)

data_sheet = json_directory('sheet_related.json')
keywords = json_directory('keywords.json')
auto_case_list = json_directory('auto_case_id.json')

def tc_location_dict():
    tc_location = load_workbook('TC_location.xlsx').active
    # stored the data in a dictionary {test_case: location}
    return {TCID: location for (TCID, location) in tc_location.iter_rows(
        max_col=2, values_only=True) if TCID is not None}

# Determine the phone type (iPhone or Android)
def phone_type(cell_data):
    iphone = keywords['iphone']
    android = keywords['android']
    phone_requirement = [0, 0]
    for cell in cell_data[pre_index:pre_index+2]:
        if matcher_split(iphone, cell):
            phone_requirement[0] = 1
        if matcher_slice(android, cell):
            phone_requirement[1] = 1
    if phone_requirement == [1, 0]:
        cell_data.append('iPhone')
    elif phone_requirement == [0, 1]:
        cell_data.append('Android')
    elif phone_requirement == [1, 1]:
        cell_data.append('Both')
    else:
        cell_data.append(' ')

# Determine the sign-in or -out
def sign_status(cell_data):
    sign_out = keywords['sign_out']
    sign_in = keywords['sign_in']
    if matcher_slice(sign_out, cell_data[pre_index]):
        cell_data.append('sign_out')
    elif matcher_slice(sign_in, cell_data[pre_index]):
        cell_data.append('sign_in')
    else:
        cell_data.append(' ')

def connection(cell_data):
    offline = keywords['offline']
    if matcher_split(offline, cell_data[pre_index]):
        cell_data.append('Offline')
    else:
        cell_data.append('Online')

def user(cell_data):
    guest = keywords['guest']
    non_guest = keywords['non_guest']
    others = keywords['others']
    primary = keywords['primary']
    if (matcher_split(guest, cell_data[pre_index]) or matcher_split(guest, cell_data[pre_index+3])) and matcher_slice(non_guest, cell_data[pre_index]) is False:
        cell_data.append('Guest')
    elif matcher_slice(others, cell_data[pre_index]) or matcher_slice(non_guest, cell_data[pre_index]) or matcher_slice(others, cell_data[pre_index+3]):
        cell_data.append('Others')
    elif matcher_split(guest, cell_data[pre_index+1]) and (matcher_slice(others, cell_data[pre_index+1]) or matcher_split(primary, cell_data[pre_index+1])):
        cell_data.append('multiple')
    else:
        cell_data.append('Driver')