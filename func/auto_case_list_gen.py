from openpyxl import load_workbook, Workbook

def formater(text):
    text = text.replace('\n', ' ')
    text = text.replace('”', '"')
    text = text.replace('“', '"')
    text = text.replace('/', '')
    text = text.split('"')
    if len(text) == 1:
        return text[0]
    else:
        return text[-2]


def auto_case_list_gen(input_name):
    output_name = input_name[:3] + '_auto_case.xlsx'

    # Loading the full-list file
    full_list = load_workbook(input_name)['auto']

    # Creating the automation-related cases spreadsheet
    wb = Workbook()
    sheet_name = ['Online_In', 'Offline_In',
                  'Online_Out', 'Offline_Out', 'Guest']
    for name in sheet_name:
        wb.create_sheet(name)
    wb.save(output_name)

    for case in full_list.iter_rows(max_col=14, values_only=True):
        if not case[0] == None and not case[0] == 'Original GM TC ID':
            case_id = case[0]

            test_step = case[7]

            command = formater(test_step)

            user = case[11]
            connection = case[12]
            sign_status = case[13]
            row = [case_id, test_step, command]
            if user == 'Driver':
                if connection == 'Online' and sign_status == 'sign_in':
                    wb['Online_In'].append(row)
                elif connection == 'Offline' and sign_status == 'sign_in':
                    wb['Offline_In'].append(row)
                elif connection == 'Online' and sign_status == 'sign_out':
                    wb['Online_Out'].append(row)
                elif connection == 'Offline' and sign_status == 'sign_out':
                    wb['Offline_Out'].append(row)
                wb.save(output_name)
            elif user == 'Guest':
                wb['Guest'].append(row)
                wb.save(output_name)

if __name__ == '__main__':
    auto_case_list_gen('W18_Production_sorted.xlsx')
