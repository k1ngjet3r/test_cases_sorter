from openpyxl import load_workbook
from openpyxl import Workbook
import re

# Defining the word matcher function


def matcher_slice(keywords, row):
    for i in range(1, 4):
        sen = (row[i].value).lower()
        for key in keywords:
            if re.search(key, sen):
                return True
    return False

# Matcher_2 matching the word by spliting the string on whitespace


def matcher_split(keywords, row):
    for i in range(1, 4):
        # Remove symbols from the lowercased sentance by replace it with a whitespace
        clean_sentance = re.sub(r'[^\w]', ' ', (row[i].value).lower())
        # Split the clean_sentance on whitespace and make it into a list
        word_list = clean_sentance.split()
        for key in keywords:
            if key in word_list:
                return True
    return False


# Load the unsorted cases file
workbook = load_workbook('Taipei_CaseList.xlsx')
sheet = workbook.active
# Saving the titles of every column in to a list and delete it for categories determination
titles = []
for title in sheet[1]:
    titles.append(title.value)
sheet.delete_rows(1)

# Create the new file named sorted cases
wb = Workbook()
sorted_cases = wb.active
# Adding worksheets for different categories
categories = ['Flash User', 'Multi User', 'Button', 'Call(Sign Out)', 'CallSMS', 'Media Center',
              'Navigation', 'HVAC', 'Invalid Cases', 'Others']
for category in categories:
    wb.create_sheet(category, int(categories.index(category)))

# keyword for layer 1
flash_user = ['flash']
mulit_user = ['multi', 'primary' 'secondary']
press_button = ['long press', 'short press', 'press "end" key']
guest = ['guest']
invalid = ['audiobook']

# Layer 2
sign_out = ["user is signed out",
            "signout the google account", "sign out the google account"]

online = ['online', 'internet']

# keyword for determining the function-related categories
navigation = ['navigat', 'go to', 'add stop', 'guidance',
              'how far', 'take me', 'address', 'traffic', 'add stop']

call_SMS = ['call', 'phone', 'message',
            'reply', 'text', 'sms', 'dial', 'Send', 'dail']

media = ['play', 'pause', 'next', 'previous',
         'volume', 'music', 'am', 'fm', 'radio', 'news', 'tune', 'tuned', 'plays', 'bluetooth', 'station', 'album', 'podcast', 'pauses', 'media', 'songs', 'playback', 'bt']


ac = ['a/c', 'temperature', 'climate', 'defroster', 'air', 'fan']


# iterate through the row
for row in sheet.rows:
    # Since the cell cannot to copy to other sheet directly, we have to store the data in a list and transfer to other sheet
    cell_data = []
    for cell in row:
        cell_data.append(cell.value)

    # Filter out the flash-, mulit-user-, press-, invalid-related cases
    if matcher_split(flash_user, row):
        wb['Flash User'].append(cell_data)

    elif matcher_split(mulit_user, row):
        wb['Multi User'].append(cell_data)

    elif matcher_slice(press_button, row):
        wb['Button'].append(cell_data)

    elif matcher_split(invalid, row):
        wb['Invalid Cases'].append(cell_data)

    # for guest user cases
    elif matcher_split(guest, row):
        continue
        # # For buttom press related cases
        # if matcher_slice(press_button, row[2]) == True:
        #     wb['Button'].append(cell_data)

        # # For callsms-related cases
        # elif (matcher_split(call_SMS, row[1]) == True or matcher_split(call_SMS, row[2]) == True or matcher_split(call_SMS, row[3])) and matcher_slice(sign_out, row[1]) != True:
        #     wb['CallSMS'].append(cell_data)

        # # For call(sign out)-related cases
        # elif (matcher_split(call_SMS, row[1]) == True or matcher_split(call_SMS, row[2]) == True) and matcher_slice(sign_out, row[1]) == True:
        #     wb['Call(Sign Out)'].append(cell_data)

        # # For HVAC-related cases
        # elif (matcher_split(ac, row[1]) == True or matcher_split(ac, row[2]) == True):
        #     wb['HVAC'].append(cell_data)

        # # For media-related cases
        # elif (matcher_split(media, row[1]) == True or matcher_split(media, row[2]) == True or matcher_split(media, row[3])) and matcher_slice(sign_out, row[1]) != True:
        #     wb["Media Center"].append(cell_data)

        # # For navigation-related cases
        # elif (matcher_slice(navigation, row[1]) == True or matcher_slice(navigation, row[2]) == True or matcher_slice(navigation, row[3])):
        #     wb['Navigation'].append(cell_data)

        # # Determine the invalid cases
        # elif matcher_slice(invalid, row[1]) == True or matcher_slice(invalid, row[2]) == True or matcher_slice(invalid, row[3]) == True:
        #     wb['Invalid cases'].append(cell_data)

        # else:
        #     wb['Others'].append(cell_data)

        # Save the file with the name
wb.save('Sorted_cases_W45.xlsx')
print('DONE')
