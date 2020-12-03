from openpyxl import load_workbook
from openpyxl import Workbook
import re

sheet_names = [
    'Difficult_cases',
    'Driver_Online_In', 'Driver_Online_Out', 'Driver_Offline_In', 'Driver_Offline_Out',
    'Guest_Online_In', 'Guest_Online_Out', 'Guest_Offline_In', 'Guest_Offline_Out',
    'Other']

difficult_cases = []


def matcher_slice(keywords, cell_data):
    sen = cell_data.lower()
    for key in keywords:
        if re.search(key, sen):
            return True
    return False


def matcher_split(keywords, cell_data):
    clean_sentance = re.sub(r'[^\w]', ' ', cell_data.lower())
    word_list = clean_sentance.split()
    for key in keywords:
        if key in word_list:
            return True
    return False


class Tc_sorter:
    def __init__(self, test_case_list, output_name, last_week_result):
        self.test_case_list = str(test_case_list)
        self.output_name = str(output_name)
        self.last_week_result = str(last_week_result)
        self.sheet = (load_workbook(self.test_case_list)).active
        self.last_week_result.sheet = (
            load_workbook(self.last_week_result)).active
        self.wb = Workbook()
        self.wb.active
        for name in sheet_names:
            self.wb.create_sheet(name, int((sheet_names).index(name)))

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def phone_type(self, cell_data):
        iphone = ['iphone', 'cp', 'wcp']
        android = ['android', 'waa', 'aa']
        phone_requirement = [0, 0]
        for cell in cell_data[1:3]:
            if matcher_split(iphone, cell):
                phone_requirement[0] = 1
            if matcher_slice(android, cell):
                phone_requirement[1] = 1
        if phone_requirement == [1, 0]:
            cell_data.append('iPhone')
        elif phone_requirement == [0, 1]:
            cell_data.append('Android')
        elif phone_requirement == [1, 1]:
            cell_data.append('Both')
        else:
            cell_data.append(' ')

    def sign_status(self, cell_data):
        sign_out = ['sign out', 'sign-out', 'signout',
                    'signed out', 'no google user is logged  in']
        if matcher_slice(sign_out, cell_data[1]):
            cell_data.append('sign_out')
        else:
            cell_data.append('sign_in')

    def connection(self, cell_data):
        offline = ['offline']
        if matcher_split(offline, cell_data[1]):
            cell_data.append('Offline')
        else:
            cell_data.append('Online')

    def formatter(self, cell_data):
        for i in range(3):
            cell_data.insert(1, '')

    def user(self, cell_data):
        guest = ['guest']
        others = ['secondary', 'user 1', 'user 2', 'user1', 'user2']
        if matcher_split(guest, cell_data[1]):
            cell_data.append('Guest')
        elif matcher_slice(others, cell_data[1]):
            cell_data.append('Others')
        else:
            cell_data.append('Driver')

    def sorting(self):
        sheet = self.sheet
        for row in sheet.iter_rows(max_col=4, values_only=True):
            cell_data = self.cell_data(row)
            self.phone_type(cell_data)
            self.user(cell_data)
            self.connection(cell_data)
            self.sign_status(cell_data)
            self.formatter(cell_data)
            # the final format will be like this:
            # ['ID', 'Pass/Fail', 'tester', 'comment', 'precondition', 'test_steps', 'expected_result', 'phone_type', 'user', 'connection', 'sign_status', 'name of tester']

            if cell_data[0] in difficult_cases:
                self.wb['Difficult_cases'].append(cell_data)

            else:
                if cell_data[8] == 'Driver' and cell_data[9] == 'Online' and cell_data[10] == 'sign_in':
                    self.wb['Driver_Online_In'].append(cell_data)
                elif cell_data[8] == 'Driver' and cell_data[9] == 'Online' and cell_data[10] == 'sign_out':
                    self.wb['Driver_Online_Out'].append(cell_data)
                elif cell_data[8] == 'Driver' and cell_data[9] == 'Offline' and cell_data[10] == 'sign_in':
                    self.wb['Driver_Offline_In'].append(cell_data)
                elif cell_data[8] == 'Driver' and cell_data[9] == 'Offline' and cell_data[10] == 'sign_out':
                    self.wb['Driver_Offline_Out'].append(cell_data)

                elif cell_data[8] == 'Guest' and cell_data[9] == 'Online' and cell_data[10] == 'sign_in':
                    self.wb['Guest_Online_In'].append(cell_data)
                elif cell_data[8] == 'Guest' and cell_data[9] == 'Online' and cell_data[10] == 'sign_out':
                    self.wb['Guest_Online_Out'].append(cell_data)
                elif cell_data[8] == 'Guest' and cell_data[9] == 'Offline' and cell_data[10] == 'sign_in':
                    self.wb['Guest_Offline_In'].append(cell_data)
                elif cell_data[8] == 'Guest' and cell_data[9] == 'Offline' and cell_data[10] == 'sign_out':
                    self.wb['Guest_Offline_Out'].append(cell_data)
                else:
                    self.wb['Other'].append(cell_data)

        self.wb.save(self.output_name)


testing = Tc_sorter('MY23_Taipei_W49.xlsx',
                    'MY23_W49.xlsx')

testing.sorting()