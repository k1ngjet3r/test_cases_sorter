from openpyxl import load_workbook
from openpyxl import Workbook, formatting, styles
import json
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from src.auto_case_list_gen import auto_case_list_gen
from src.div import *
from src.detail import *

# the index of the precondition
pre_index = 5


def json_directory(json_name):
    with open('json\\' + json_name) as f:
        return json.load(f)

# loading other list
# logan_list_sheet = load_workbook('logan_list.xlsx').active
# logan_list = [r[0] for r in logan_list_sheet.iter_rows(
#     max_col=1, max_row=335, values_only=True)]

# Loading automation case list
# auto_sheet = load_workbook('Phase_1.xlsx')['Phone_projection_1']
# auto_list = []
# for row in auto_sheet.iter_rows(max_col=1, values_only=True):
#     if row[0] != 'TCID' and row[0] != None and row[0] != '':
#         try:
#             auto_list.append(row[0].lower())
#         except:
#             break


class Tc_sorter:
    def __init__(self, test_case_list, last_week, continue_from=False):
        print('Initiallizing...')
        self.test_case_list = str(test_case_list)
        self.output_name = test_case_list[:-10] + 'sorted.xlsx'
        self.last_week = str(last_week)
        self.sheet = (load_workbook(self.test_case_list)).active
        print('{} loaded successfully'.format(self.test_case_list))

        # Loading JSON file
        self.data_sheet = json_directory('sheet_related.json')
        self.keywords = json_directory('keywords.json')
        self.auto_case_list = json_directory('auto_case_id.json')

        # Dump('W17_Main_sorted.xlsx').dump()
        self.tcid_and_sheet = json_directory('locked_tcid.json')

        # Loading the resut from last week
        self.last_week_result = (
            load_workbook(self.last_week))
        print('{} loaded successfully'.format(self.last_week))

        self.result_dv = DataValidation(
            type='list', formula1='"Pass, Fail, Hold, Invalid"', allow_blank=True)
        self.executer_dv = DataValidation(
            type='list', formula1='"maggie.chang,yvonne.chien,logan.chang,jeter.lin,jack.hsu,joan.chen,mark.mo,sarah.chiang"', allow_blank=True)

        if continue_from == False:
            self.wb = Workbook()
            # self.wb.active
            for name in self.data_sheet['sheet_names']:
                self.wb.create_sheet(
                    name, int((self.data_sheet['sheet_names']).index(name)))
                self.wb[name].append(self.data_sheet['titles'])
            for fail_name in self.data_sheet['fail_case_sheet']:
                self.wb.create_sheet(fail_name, -1)
                self.wb[fail_name].append(self.data_sheet['fail_case_titles'])
            print('Output file initiallized')

        else:
            self.wb = load_workbook(self.output_name)

    def cell_validation(self, sheetname_list, name_and_num_dict):
        num_list = [i for i in name_and_num_dict.values()]
        for name, num in zip(sheetname_list, num_list):
            if num >= 2:
                result_cell = "B2:B{}".format(num+1)
                tester_cell = "C2:C{}".format(num+1)

                result_dv = DataValidation(
                    type='list', formula1='"Pass, Fail, Hold, Invalid"', allow_blank=True)
                executer_dv = DataValidation(
                    type='list', formula1='"maggie.chang,yvonne.chien,logan.chang,jeter.lin,jack.hsu,joan.chen,mark.mo,sarah.chiang"', allow_blank=True)

                self.wb[name].add_data_validation(result_dv)
                self.wb[name].add_data_validation(executer_dv)

                result_dv.add(result_cell)
                executer_dv.add(tester_cell)

    def conditional_formatting(self, sheetname_list, name_and_num_dict):
        num_list = [i for i in name_and_num_dict.values()]
        green = 'D9EAD3'
        green_fill = styles.PatternFill(
            start_color=green, end_color=green, fill_type='solid')
        blue = 'CFE2F3'
        blue_fill = styles.PatternFill(
            start_color=blue, end_color=blue, fill_type='solid')
        red = 'F4CCCC'
        red_fill = styles.PatternFill(
            start_color=red, end_color=red, fill_type='solid')
        gray = 'CCCCCC'
        gray_fill = styles.PatternFill(
            start_color=gray, end_color=gray, fill_type='solid')

        results = ['Pass', 'Fail', 'Hold', 'Invalid']
        colors = [green_fill, red_fill, blue_fill, gray_fill]
        for name, num in zip(sheetname_list, num_list):
            if num >= 2:
                result_cell = "B2:B{}".format(num+1)
                for r, c in zip(results, colors):
                    self.wb[name].conditional_formatting.add(
                        result_cell, CellIsRule(operator='containsText', formula=[r], fill=c))
        self.wb.save(self.output_name)

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def formatter(self, cell_data):
        for _ in range(4):
            cell_data.insert(1, '')

    def last_week_result_dict(self):
        last_week_dict = {}
        last_week = self.last_week_result
        for sheet in last_week.worksheets:
            if sheet.title not in ['Fail Cases China', 'Cases need update', 'Summary']:
                for last_week_row in sheet.iter_rows(max_col=5, values_only=True):
                    last_week_cell = self.cell_data(last_week_row)
                    if last_week_cell[0] != 'Original GM TC ID':
                        last_week_dict[last_week_cell[0]] = last_week_cell[1:]
        return last_week_dict

    def generate_auto_list(self):
        auto_case_list_gen(self.output_name)

    def sorted_manually(self, cell_data, name_and_num):
        # Sorting the case based on organized sheet dict
        tcid_and_sheet = json_directory('tcid_and_sheet.json')
        for sheet_name in tcid_and_sheet:
            if cell_data[0] in tcid_and_sheet[sheet_name]:
                sheet_name = sheet_name.lower()
                self.wb[sheet_name.lower()].append(cell_data)
                name_and_num[sheet_name] += 1
                return True
        return False

    def sorting(self):
        print('Opening a new sheet...')
        sheet = self.sheet
        print('Last week result loaded successfully')
        # difficult_cases_list = self.difficult_cases()
        print('Difficult case list generated')
        location_dict = tc_location_dict()
        print('Test case location dictionary generated')
        print('Iterating through the test plan......')

        k = 1

        # counter
        # overall_num = [i for i in self.data_sheet['sheet_names']]

        name_and_num = {name: 0 for name in self.data_sheet['sheet_names']}

        # Iterate through the unprocessd test cases
        # Only getting the first 5 values of each row (tc, precondition, test_steps, expected_result, test_objective}
        for row in sheet.iter_rows(max_col=5, values_only=True):
            print('Iterate case no. {}: {}'.format(k, row[0]))
            # turn the data into a list
            cell_data = []
            for cell in row:
                if cell is not None:
                    cell_data.append(cell)
                else:
                    cell_data.append('none')

            # adding 'pass/fail', 'Tester', 'Automation_comment', 'bug ID', 'Note' to the list
            self.formatter(cell_data)
            # determine the phone type
            phone_type(cell_data)
            # determine the user type
            user(cell_data)
            # determine online/offline
            connection(cell_data)
            # determine sign-in/sign-out
            sign_status(cell_data)
            k += 1

            # finding the case's origin
            if cell_data[0] in location_dict:
                cell_data.append(location_dict[cell_data[0]])
            else:
                cell_data.append('none')

            last_week_dict = self.last_week_result_dict()

            for i in range(4):
                if cell_data[0] in last_week_dict:
                    cell_data.append(last_week_dict[cell_data[0]][i])
                else:
                    cell_data.append(' ')

            # Forming the final format
            cell_data = cell_data[:5] + [cell_data[-1]] + cell_data[5:-1]

            # Distributing the test case to the desinated sheet

            # if cell_data[0] in logan_list:
            #     self.wb['Logan'].append(cell_data)

            # Append the case to "difficult_cases" sheet based on last week's result
            if cell_data[-3] == 'Fail':
                self.wb['Difficult_cases'].append(cell_data)
                name_and_num['Difficult_cases'] += 1

            # elif cell_data[0].lower() in auto_list:
            #     self.wb['automation'].append(cell_data)
            #     num_automation += 1

            elif cell_data[0] in self.tcid_and_sheet['DID']:
                self.wb['DID'].append(cell_data)
                name_and_num['DID'] += 1

            elif cell_data[0] in self.tcid_and_sheet['User_Build']:
                self.wb['User_Build'].append(cell_data)
                name_and_num['User_Build'] += 1

            # Append the case to "auto" if the case ID is in the "auto_case_id.json"
            elif cell_data[0] in self.auto_case_list['Auto'] or cell_data[0] in self.auto_case_list['fuel_sim']:
                self.wb['Auto'].append(cell_data)
                name_and_num['Auto'] += 1

            elif did_case(cell_data):
                self.wb['DID'].append(cell_data)
                name_and_num['DID'] += 1

            elif user_build_only(cell_data):
                self.wb['User_Build'].append(cell_data)
                name_and_num['User_Build'] += 1

            elif tja(cell_data):
                self.wb['TJA'].append(cell_data)
                name_and_num['TJA'] += 1

            elif usb_update(cell_data):
                self.wb['Usb_update'].append(cell_data)
                name_and_num['Usb_update'] += 1

            elif nav_case(cell_data):
                self.wb['Nav'].append(cell_data)
                name_and_num['Nav'] += 1

            # elif self.fuel_sim(cell_data):
            #     self.wb['Fuel_sim'].append(cell_data)

            elif bench_only(cell_data):
                self.wb['Bench_Only'].append(cell_data)
                name_and_num['Bench_Only'] += 1

            elif call_SMS(cell_data):
                self.wb['Call&SMS'].append(cell_data)
                name_and_num['Call&SMS'] += 1

            elif trailer_case(cell_data):
                self.wb['Trailer'].append(cell_data)
                name_and_num['Trailer'] += 1

            elif screen_size_13(cell_data):
                self.wb['13_inch']
                name_and_num['13_inch'] += 1

            # elif self.ac_only(cell_data):
            #     self.wb['ac_only'].append(cell_data)

            else:
                i = 11
                if cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Online_In'].append(cell_data)
                    name_and_num['Driver_Online_In'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Online_Out'].append(cell_data)
                    name_and_num['Driver_Online_Out'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Offline_In'].append(cell_data)
                    name_and_num['Driver_Offline_In'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Offline_Out'].append(cell_data)
                    name_and_num['Driver_Offline_Out'] += 1
                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Guest_Online_In'].append(cell_data)
                    name_and_num['Guest_Online_In'] += 1
                else:
                    self.wb['Other'].append(cell_data)
                    name_and_num['Other'] += 1

        print('Saving the file named {}\n'.format(self.output_name))
        self.wb.save(self.output_name)

        print('===============================================')
        print('[SUMMARY]')
        overall = 0
        for name in name_and_num:
            print('num_{}: {}'.format(name, name_and_num[name]))
            overall += int(name_and_num[name])

        print('Overall Sum: {}'.format(overall))
        print('===============================================')

        print('Adding data validation to output')
        self.cell_validation(self.data_sheet['sheet_names'], name_and_num)
        self.wb.save(self.output_name)

        print('Conditional Formatting the cell')
        self.conditional_formatting(
            self.data_sheet['sheet_names'], name_and_num)

        print('Generating automation case list...')
        self.generate_auto_list()

        print('Done')


if __name__ == '__main__':
    # __init__(self, test_case_list, last_week, continue_from=False)
    testing = Tc_sorter('W19_production_cases.xlsx',
                        'W18_Production_sorted.xlsx', continue_from=False)
    testing.sorting()

    testing = Tc_sorter('W19_Main_cases.xlsx',
                        'W18_Main_sorted.xlsx', continue_from=False)
    testing.sorting()
