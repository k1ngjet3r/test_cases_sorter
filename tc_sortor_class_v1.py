from openpyxl import load_workbook
from openpyxl import Workbook
import re

workbook = load_workbook('MY22_1499s.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(max_col=5, values_only=False):
    print([cell.value for cell in row])
