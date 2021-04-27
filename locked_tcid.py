from openpyxl import load_workbook
import json, shutil


class Dump:
    def __init__(self, file_name):
        self.wb = load_workbook(file_name)

    def intel(self, sheet_list):
        intel = {}
        for sheet in sheet_list:
            print('generate the {} list'.format(sheet))
            temp = []
            for row in self.wb[sheet].iter_rows(max_col=1, values_only=True):
                try:
                    temp.append(row[0].lower())
                except:
                    break
            intel[sheet] = temp[1:]
        return intel

    def dump(self, sheet_list):
        with open('json\\locked_tcid.json', 'w') as outfile:
            json.dump(self.intel(sheet_list), outfile)

    def update(self, sheet_list):
        with open('json\\locked_tcid.json', 'r') as infile:
            json_file = json.load(infile)
            tcid_dict = self.intel(sheet_list)
            for tcid in tcid_dict:
                json_file[tcid] = tcid_dict[tcid]

        with open('json\\locked_tcid.json', 'w') as outfile:
            json.dump(json_file, outfile)



if __name__ == '__main__':
    Dump('W17_Main_sorted.xlsx').dump(['DID', 'User_Build'])
    Dump('W17_Production_sorted.xlsx').update(['Security Lock', 'Google Setup wizard'])
