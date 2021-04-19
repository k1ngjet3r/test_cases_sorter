from openpyxl import load_workbook
from openpyxl import Workbook, formatting, styles
import re
import json
from matcher.matcher import matcher_split, matcher_slice
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from auto_case_list_gen import auto_case_list_gen

# the index of the precondition
pre_index = 5


def json_directory(json_name):
    with open('json\\' + json_name) as f:
        return json.load(f)

# loading other list
# logan_list_sheet = load_workbook('logan_list.xlsx').active
# logan_list = [r[0] for r in logan_list_sheet.iter_rows(
#     max_col=1, max_row=335, values_only=True)]

# Loading automation case list
# auto_sheet = load_workbook('Phase_1.xlsx')['Phone_projection_1']
# auto_list = []
# for row in auto_sheet.iter_rows(max_col=1, values_only=True):
#     if row[0] != 'TCID' and row[0] != None and row[0] != '':
#         try:
#             auto_list.append(row[0].lower())
#         except:
#             break

class Tc_sorter:
    def __init__(self, test_case_list, last_week, continue_from=False):
        print('Initiallizing...')
        self.test_case_list = str(test_case_list)
        self.output_name = test_case_list[:3] + 'Main_sorted.xlsx'
        self.last_week = str(last_week)
        self.sheet = (load_workbook(self.test_case_list)).active
        print('{} loaded successfully'.format(self.test_case_list))

        # Loading JSON file
        self.data_sheet = json_directory('sheet_related.json')
        self.keywords = json_directory('keywords.json')
        self.auto_case_list = json_directory('auto_case_id.json')

        # Loading the resut from last week
        self.last_week_result = (
            load_workbook(self.last_week))
        print('{} loaded successfully'.format(self.last_week))

        self.result_dv = DataValidation(
            type='list', formula1='"Pass, Fail, Hold, Invalid"', allow_blank=True)
        self.executer_dv = DataValidation(
            type='list', formula1='"maggie.chang,yvonne.chien,logan.chang,jeter.lin,jack.hsu,joan.chen,mark.mo,sarah.chiang"', allow_blank=True)

        if continue_from == False:
            self.wb = Workbook()
            # self.wb.active
            for name in self.data_sheet['sheet_names']:
                self.wb.create_sheet(
                    name, int((self.data_sheet['sheet_names']).index(name)))
                self.wb[name].append(self.data_sheet['titles'])
            for fail_name in self.data_sheet['fail_case_sheet']:
                self.wb.create_sheet(fail_name, -1)
                self.wb[fail_name].append(self.data_sheet['fail_case_titles'])
            print('Output file initiallized')

        else:
            self.wb = load_workbook(self.output_name)

    def cell_validation(self, sheetname_list, num_list):
        for name, num in zip(sheetname_list, num_list):
            if num >= 2:
                result_cell = "B2:B{}".format(num+1)
                tester_cell = "C2:C{}".format(num+1)

                result_dv = DataValidation(
                    type='list', formula1='"Pass, Fail, Hold, Invalid"', allow_blank=True)
                executer_dv = DataValidation(
                    type='list', formula1='"maggie.chang,yvonne.chien,logan.chang,jeter.lin,jack.hsu,joan.chen,mark.mo,sarah.chiang"', allow_blank=True)

                self.wb[name].add_data_validation(result_dv)
                self.wb[name].add_data_validation(executer_dv)

                result_dv.add(result_cell)
                executer_dv.add(tester_cell)

    def conditional_formatting(self, sheetname_list, num_list):
        green = 'D9EAD3'
        green_fill = styles.PatternFill(
            start_color=green, end_color=green, fill_type='solid')
        blue = 'CFE2F3'
        blue_fill = styles.PatternFill(
            start_color=blue, end_color=blue, fill_type='solid')
        red = 'F4CCCC'
        red_fill = styles.PatternFill(
            start_color=red, end_color=red, fill_type='solid')
        gray = 'CCCCCC'
        gray_fill = styles.PatternFill(
            start_color=gray, end_color=gray, fill_type='solid')

        results = ['Pass', 'Fail', 'Hold', 'Invalid']
        colors = [green_fill, red_fill, blue_fill, gray_fill]
        for name, num in zip(sheetname_list, num_list):
            if num >= 2:
                result_cell = "B2:B{}".format(num+1)
                for r, c in zip(results, colors):
                    self.wb[name].conditional_formatting.add(
                        result_cell, CellIsRule(operator='containsText', formula=[r], fill=c))
        self.wb.save(self.output_name)

    def Automation_cases(self):
        auto_file = load_workbook('automation_cases.xlsx').active
        return [tcid[0] for tcid in auto_file.iter_rows(max_col=1, values_only=True)]

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def phone_type(self, cell_data):
        iphone = self.keywords['iphone']
        android = self.keywords['android']
        phone_requirement = [0, 0]
        for cell in cell_data[pre_index:pre_index+2]:
            if matcher_split(iphone, cell):
                phone_requirement[0] = 1
            if matcher_slice(android, cell):
                phone_requirement[1] = 1
        if phone_requirement == [1, 0]:
            cell_data.append('iPhone')
        elif phone_requirement == [0, 1]:
            cell_data.append('Android')
        elif phone_requirement == [1, 1]:
            cell_data.append('Both')
        else:
            cell_data.append(' ')

    def sign_status(self, cell_data):
        sign_out = self.keywords['sign_out']
        if matcher_slice(sign_out, cell_data[pre_index]):
            cell_data.append('sign_out')
        else:
            cell_data.append('sign_in')

    def connection(self, cell_data):
        offline = self.keywords['offline']
        if matcher_split(offline, cell_data[pre_index]):
            cell_data.append('Offline')
        else:
            cell_data.append('Online')

    def formatter(self, cell_data):
        for _ in range(4):
            cell_data.insert(1, '')

    def user(self, cell_data):
        guest = self.keywords['guest']
        non_guest = self.keywords['non_guest']
        others = self.keywords['others']
        primary = self.keywords['primary']
        if (matcher_split(guest, cell_data[pre_index]) or matcher_split(guest, cell_data[pre_index+3])) and matcher_slice(non_guest, cell_data[pre_index]) is False:
            cell_data.append('Guest')
        elif matcher_slice(others, cell_data[pre_index]) or matcher_slice(non_guest, cell_data[pre_index]) or matcher_slice(others, cell_data[pre_index+3]):
            cell_data.append('Others')
        elif matcher_split(guest, cell_data[pre_index+1]) and (matcher_slice(others, cell_data[pre_index+1]) or matcher_split(primary, cell_data[pre_index+1])):
            cell_data.append('multiple')
        else:
            cell_data.append('Driver')

    def bench_only(self, cell_data):
        press_button = self.keywords['push_button']
        cluster = self.keywords['cluster']
        speed_limit = self.keywords['speed_limit']
        expection = self.keywords['expection']
        for cell in cell_data[pre_index+1:pre_index+5]:
            if (matcher_slice(press_button, cell) or matcher_slice(cluster, cell) or matcher_slice(speed_limit, cell)) and not matcher_slice(expection, cell):
                return True
            return False

    def ac_only(self, cell_data):
        ac = self.keywords['ac']
        ac_split = self.keywords['ac_split']
        for cell in cell_data[pre_index:pre_index+3]:
            if matcher_slice(ac, cell) or matcher_split(ac_split, cell):
                return True
            return False

    def tc_location_dict(self):
        tc_location = load_workbook('TC_location.xlsx').active
        # stored the data in a dictionary {test_case: location}
        return {TCID: location for (TCID, location) in tc_location.iter_rows(
            max_col=2, values_only=True) if TCID is not None}

    def last_week_result_dict(self):
        last_week_dict = {}
        last_week = self.last_week_result
        for sheet in last_week.worksheets:
            if sheet.title not in ['Fail Cases China', 'Cases need update', 'Summary']:
                for last_week_row in sheet.iter_rows(max_col=5, values_only=True):
                    last_week_cell = self.cell_data(last_week_row)
                    if last_week_cell[0] != 'Original GM TC ID':
                        last_week_dict[last_week_cell[0]] = last_week_cell[1:]
        return last_week_dict

    def nav_case(self, cell_data):
        # Finding the navigation-related cases using TCID
        # Formatting the TCID
        tcid = [i.lower() for i in cell_data[0].split('_')]
        if 'maps' in tcid:
            return True
        return False

    def call_SMS(self, cell_data):
        callsms = self.keywords['call_sms']
        if matcher_slice(callsms, cell_data[pre_index+1]):
            return True
        return False

    def fuel_sim(self, cell_data):
        fuel = self.keywords['fuel_sim']
        if matcher_slice(fuel, cell_data[pre_index+1]):
            return True
        return False

    def did_case(self, cell_data):
        did = self.keywords['did']
        user = self.keywords['user']
        # search DID-related case ID in test obnjective
        if matcher_slice(did, cell_data[pre_index+3]) and not matcher_slice(user, cell_data[pre_index+3]):
            return True
        return False

    def user_build_only(self, cell_data):
        user = self.keywords['user']
        if matcher_slice(user, cell_data[pre_index+3]) or matcher_slice(user, cell_data[pre_index]):
            return True
        return False

    def screen_size_13(self, cell_data):
        thirdteen_inch = self.keywords['13_inch']
        for i in range(4):
            if matcher_slice(thirdteen_inch, cell_data[pre_index+i]):
                return True
        return False

    def trailer_case(self, cell_data):
        trailer_kw = self.keywords['trailer']
        for i in range(4):
            if matcher_slice(trailer_kw, cell_data[pre_index+i]):
                return True
        return False

    def generate_auto_list(self):
        auto_case_list_gen(self.output_name)

    def sorting(self):
        print('Opening a new sheet...')
        sheet = self.sheet
        print('Last week result loaded successfully')
        # difficult_cases_list = self.difficult_cases()
        print('Difficult case list generated')
        location_dict = self.tc_location_dict()
        print('Test case location dictionary generated')
        print('Iterating through the test plan......')

        k = 1

        # counter
        
        # num_diff = 0
        # num_ben = 0
        # num_dri_on_in = 0
        # num_dri_on_out = 0
        # num_dri_off_in = 0
        # num_dri_off_out = 0
        # num_ges_on_in = 0
        # num_other = 0
        # num_nav = 0
        # num_auto = 0
        # num_callsms = 0
        # num_did = 0
        # num_user_build = 0
        # num_13_inch = 0
        # num_trailer = 0
        # num_automation = 0
        overall_num = ['num_diff', 'num_ben', 'num_dri_on_in', 'num_dri_on_out', 'num_dri_off_in', 'num_dri_off_out',
                       'num_ges_on_in', 'num_other', 'num_nav', 'num_auto', 'num_callsms', 'num_did', 'num_user_build', 'num_13_inch', 'num_trailer', 'num_automation']
        
        name_and_num = {name: 0 for name in overall_num}


        # Iterate through the unprocessd test cases
        # Only getting the first 5 values of each row (tc, precondition, test_steps, expected_result, test_objective}
        for row in sheet.iter_rows(max_col=5, values_only=True):
            print('Iterate case no. {}'.format(k))
            # turn the data into a list
            cell_data = []
            for cell in row:
                if cell is not None:
                    cell_data.append(cell)
                else:
                    cell_data.append('none')

            # adding 'pass/fail', 'Tester', 'Automation_comment', 'bug ID', 'Note' to the list
            self.formatter(cell_data)
            # determine the phone type
            self.phone_type(cell_data)
            # determine the user type
            self.user(cell_data)
            # determine online/offline
            self.connection(cell_data)
            # determine sign-in/sign-out
            self.sign_status(cell_data)
            k += 1

            if cell_data[0] in location_dict:
                cell_data.append(location_dict[cell_data[0]])
            else:
                cell_data.append('none')

            last_week_dict = self.last_week_result_dict()

            for i in range(4):
                if cell_data[0] in last_week_dict:
                    cell_data.append(last_week_dict[cell_data[0]][i])
                else:
                    continue

            # Forming the final format
            cell_data = cell_data[:5] + [cell_data[-1]] + cell_data[5:-1]

            # Distributing the test case to the desinated sheet

            # Append the case to "difficult_cases" sheet based on last week's result
            # if cell_data[0] in logan_list:
            #     self.wb['Logan'].append(cell_data)

            if cell_data[-3] == 'Fail':
                self.wb['Difficult_cases'].append(cell_data)
                name_and_num['num_diff'] += 1

            # elif cell_data[0].lower() in auto_list:
            #     self.wb['automation'].append(cell_data)
            #     num_automation += 1

            # Append the case to "auto" if the case ID is in the "auto_case_id.json"
            elif cell_data[0] in self.auto_case_list['auto'] or cell_data[0] in self.auto_case_list['fuel_sim']:
                self.wb['auto'].append(cell_data)
                name_and_num['num_auto'] += 1

            elif self.did_case(cell_data):
                self.wb['DID'].append(cell_data)
                name_and_num['num_did'] += 1

            elif self.user_build_only(cell_data):
                self.wb['User_Build'].append(cell_data)
                name_and_num['num_user_build'] += 1

            elif self.nav_case(cell_data):
                self.wb['Nav'].append(cell_data)
                name_and_num['num_nav'] += 1

            # elif self.fuel_sim(cell_data):
            #     self.wb['Fuel_sim'].append(cell_data)

            elif self.bench_only(cell_data):
                self.wb['Bench_only'].append(cell_data)
                name_and_num['num_ben'] += 1

            elif self.call_SMS(cell_data):
                self.wb['Call&SMS'].append(cell_data)
                name_and_num['num_callsms'] += 1
            
            elif self.trailer_case(cell_data):
                self.wb['trailer'].append(cell_data)
                name_and_num['num_trailer'] += 1

            elif self.screen_size_13(cell_data):
                self.wb['13_inch']
                name_and_num['num_13_inch'] += 1

            # elif self.ac_only(cell_data):
            #     self.wb['ac_only'].append(cell_data)

            else:
                i = 11
                if cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Online_In'].append(cell_data)
                    name_and_num['num_dri_on_in'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Online_Out'].append(cell_data)
                    name_and_num['num_dri_on_out'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Offline_In'].append(cell_data)
                    name_and_num['num_dri_off_in'] += 1
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Offline_Out'].append(cell_data)
                    name_and_num['num_dri_off_out'] += 1

                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Guest_Online_In'].append(cell_data)
                    name_and_num['num_ges_on_in'] += 1

                else:
                    self.wb['Other'].append(cell_data)
                    name_and_num['num_other'] += 1

        print('Saving the file named {}\n'.format(self.output_name))
        self.wb.save(self.output_name)
        print('===============================================')
        print('[SUMMARY]')
        overall = 0
        for name in name_and_num:
            print('{}: {}'.format(name, name_and_num[name]))
            overall += int(name_and_num[name])
        
        print('Overall: {}'.format(overall))

                       
    

        print('Adding data validation to output')
        self.cell_validation(self.data_sheet['sheet_names'], overall_num)
        self.wb.save(self.output_name)

        print('Conditional Formatting the cell')
        self.conditional_formatting(self.data_sheet['sheet_names'], overall_num)

        print('Generating automation case list...')
        self.generate_auto_list()

        print('Done')

if __name__ == '__main__':
    # __init__(self, test_case_list, last_week, continue_from=False)
    testing = Tc_sorter('W17_358_MainLine_cases.xlsx', 'Logan不要拉拉拉拉拉_W16_Main_sorted.xlsx', continue_from=False)
    testing.sorting()
  