from openpyxl import load_workbook
import json

wb = load_workbook('TC_location.xlsx').active

location_dict = {TCID: location for (TCID, location) in wb.iter_rows(
    max_col=2, values_only=True) if TCID is not None}

with open('json\\tc_location.json', 'w') as outfile:
    json.dump(location_dict, outfile)
