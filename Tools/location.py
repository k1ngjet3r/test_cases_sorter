from openpyxl import load_workbook

tc_location = load_workbook('TC_location.xlsx').active

tc_location_dict = {i: l for (i, l) in tc_location.iter_rows(
    max_col=2, values_only=True) if i is not None}

print(tc_location_dict)
