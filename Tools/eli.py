from openpyxl import Workbook, load_workbook
import json


def json_directory(json_name):
    with open('json/' + json_name) as f:
        return json.load(f)


class eliminator:
    def __init__(self, wb, full_cases, tcid_only=True):
        # self.name = wb
        self.partial_list = load_workbook(wb)['Packages_71']
        # self.partial_list = self.wb['Packages_71']
        self.full_list = load_workbook(full_cases).active
        if tcid_only == True:
            self.max_column = 1
        else:
            self.max_column = 5

    def differentiator(self):
        wb = Workbook()
        wb.create_sheet('intersection')
        wb.create_sheet('symmetric difference')

        partial_tcid = [row[0].lower() for row in self.partial_list.iter_rows(
            max_col=1, max_row=72, values_only=True)]

        # partial_tcid = []

        # for tcid in self.partial_list.iter_rows(max_col=1, values_only=True):
        #     try:
        #         partial_tcid.append(row[0].lower())
        #     except:
        # break
        print('Num partial list: {}'.format(len(partial_tcid)))

        for row in self.full_list.iter_rows(max_col=self.max_column, values_only=True):
            row_data = [i for i in row]
            if row_data[1] in partial_tcid:
                wb['intersection'].append(row_data)
            else:
                wb['symmetric difference'].append(row_data)

        wb.save('cases.xlsx')

    def differentiator_counter(self):
        current = 1
        matched = 0
        not_matched = 0
        partial_tcid = [row[0] for row in self.partial_list.iter_rows(
            max_col=1, values_only=True)]

        for row in self.full_list.iter_rows(max_col=1, values_only=True):
            print('iterate case number {}'.format(current))
            try:
                current += 1
                if row[0] in partial_tcid:
                    matched += 1
                else:
                    not_matched += 1
            except:
                break
        print('Matched: {}'.format(matched))
        print('Not Matched: {}'.format(not_matched))

    def differentiator_same_sheet(self):
        self.wb.create_sheet('Production_with_script_detail')
        current = 1
        matched = 0

        partial_tcid = [row[0].lower() for row in self.partial_list.iter_rows(
            max_col=1, values_only=True)]

        for row in self.full_list.iter_rows(max_col=1, values_only=True):
            try:
                print('iterate case number {}'.format(current))
                current += 1
                if row[0].lower() in partial_tcid:
                    matched += 1
                    self.wb['Production_with_script'].append(row)
                    self.wb.save(self.name)
            except:
                break
        # self.wb.save(self.name)
        print('Matched: {}'.format(matched))

    def differentiator_with_sorted(self, sorted_sheet):
        sorted_wb = load_workbook(sorted_sheet)

        wb = Workbook()
        wb.create_sheet('intersection')
        wb.create_sheet('symmetric difference')

        sheet_names = json_directory('sheet_related.json')['sheet_names']

        partial_tcid = partial_tcid = [row[0].lower() for row in self.partial_list.iter_rows(
            max_col=1, max_row=72, values_only=True)]

        for name in sheet_names:
            try:
                for case in sorted_wb[name].iter_rows(max_col=1, values_only=True):
                    try:
                        if case[0].lower() in partial_tcid:
                            wb['intersection'].append(row_data)
                    except:
                        break
            except:
                continue
        wb.save('cases.xlsx')


if __name__ == '__main__':
    eliminator('Side_Quest.xlsx', 'MY22_signed_intersect_cases.xlsx',
               tcid_only=False).differentiator_with_sorted('W17_Production_sorted.xlsx')
