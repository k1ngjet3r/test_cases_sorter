from re import match
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

class pkg_ctrl:
    def __init__(self, package_design, case_list, output_name, sheetname=False):
        if sheetname == True:
            self.package_design = load_workbook(package_design)['Function Package Design']
        else:
            self.package_design = load_workbook(package_design).active
        self.case_list = load_workbook(case_list).active
        self.output_file = Workbook()
        self.output_name = output_name
        self.sheetname = []
    
    def case_list_2_list(self):
        id_list = []
        for id in self.case_list.iter_rows(max_col=1, values_only=True):
            if id[0] != None or id[0] != '':
                id_list.append(id[0].lower())
        return id_list

    def comparision(self):
        wb = self.output_file.active
        id_list = self.case_list_2_list()
        total = 0
        matched = 0
        not_matched = 0
        for case in self.package_design.iter_rows(max_col=6, values_only=True):
            total += 1
            print('Iterating case number {}'.format(total))
            if case[1].lower() in id_list:
                matched += 1
                print('Matched count: {}'.format(matched))
                wb.append(case)
            else:
                not_matched += 1
            self.output_file.save(self.output_name)

        print('[Summary]')
        print('Iterate through {} cases'.format(total))
        print('Matched cases: {}'.format(matched))
        print('Not match: {}'.format(not_matched))

if __name__ == '__main__':
    pkg = pkg_ctrl('Package-Design-MY22-GB10-TT.xlsx', 'MY22_signed_intersect_cases.xlsx', 'MY22_signed_package.xlsx',sheetname=True)
    pkg.comparision()