from openpyxl import load_workbook

wb = load_workbook('Phase_1.xlsx')['Phone_projection_1']

precode_list = []

for i in wb.iter_rows(max_col=2, values_only=True):
    if i[0] != 'TCID':
        try:
            for j in i[-1].split(','):
                if j not in precode_list:
                    precode_list.append(j)
        except AttributeError:
            break


print(len(precode_list))

s = str(precode_list)

s = s.replace('[','')
s = s.replace(']','')
s = s.replace("'", '')
s = s.replace(', ', '\n')
print(s)