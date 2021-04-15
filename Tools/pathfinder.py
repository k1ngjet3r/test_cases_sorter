import os
from shutil import copy2
from openpyxl import load_workbook



def path_finder(filename, search_path):
    for root, direction, files in os.walk(search_path):
        if filename.lower() in files:
            return root+'\\'+filename

def move_to_destination(target):
    destination = 'C:\\Users\\GM-PC-03\\Desktop\\phase_1\\Wireless_Android_auto_1\\'
    copy2(target, destination)


if __name__ == '__main__':
    case_list = load_workbook('Phase_1.xlsx')['Wireless_Android_auto_1']
    search_path = 'C:\\Users\\GM-PC-03\\Desktop\\Automation\\Gerrit\\src\\1080p_tt\\scripts\\bj'
    current = 1
    found = 0
    not_found = 0
    for case in case_list.iter_rows(max_col=1, values_only=True):
        try:
            print('finding case number: {}'.format(current))
            current += 1
            
            case_name = case[0] + '.xml'
            if path_finder(case_name, search_path):
                print('found')
                found += 1
                move_to_destination(path_finder(case_name, search_path))
            else:
                not_found += 1
        except TypeError:
            break
        
    print('[SUMMARY]')
    print('Found: {}'.format(found))
    print('Not found: {}'.format(not_found))