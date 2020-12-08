from openpyxl import load_workbook
from openpyxl import Workbook

file_names = []

# create the output file
past_result = Workbook()
past_result.active

# read the test case 1499 list
case_list = load_workbook('records/MY22_1499_list.xlsx').active
for id in case_list.iter_rows(max_col=1, values_only=True):
    print(id)
