from openpyxl import load_workbook
import json, shutil


class Dump:
    def __init__(self, file_name):
        self.wb = load_workbook(file_name)

    def intel(self):
        target_sheet = ['DID', 'User_Build']
        intel = {}
        for sheet in target_sheet:
            print('generate the {} list'.format(sheet))
            temp = []
            for row in self.wb[sheet].iter_rows(max_col=1, values_only=True):
                try:
                    temp.append(row[0].lower())
                except:
                    break
            intel[sheet] = temp[1:]
        return intel

    def dump(self):
        with open('json\\tcid_and_sheet.json', 'w') as outfile:
            json.dump(self.intel(), outfile)

        shutil.move('C:\\Users\\GM-PC-03\\Documents\\Python\\k1ngjet3r\\test_cases_sorter\\tcid_and_sheet.json', 'C:\\Users\\GM-PC-03\\Documents\\Python\\k1ngjet3r\\test_cases_sorter\\json\\tcid_and_sheet.json')


if __name__ == '__main__':
    Dump('W17_Main_sorted.xlsx').dump()
