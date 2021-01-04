from openpyxl import load_workbook
from openpyxl import Workbook
import re

sheet_names = [
    'Difficult_cases', 'Bench_only', 'ac_only',
    'Driver_Online_In', 'Driver_Online_Out', 'Driver_Offline_In', 'Driver_Offline_Out',
    'Guest_Online_In', 'Other']

fail_case_sheet = ['Fail Cases Warren',
                   'Fail Cases China', 'Cases for Lui Fei']

fail_case_title = ['Date of failure', 'Ticket Filed', 'Original GM TC ID', 'Product Line', 'Case Location', 'Result Taipei', 'BUG ID',
                   'Precondition', 'Test steps', 'Expected', 'Automation Comment', 'Result Beijing, Nanjing, Warren', 'Comment Beijing, Nanjing, Warren', 'Tester']

titles = ['Original GM TC ID', 'Pass/Fail', 'Tester', 'Automation Comment', 'Bug ID', 'Note',
          'Precondition', 'Test steps', 'Expected', 'Testing Objective', 'Phone', 'User', 'Online/Offline', 'Sign Status', 'Location',
          'W50_result', 'W50_tester', 'W50_Automation_Comment']


def matcher_slice(keywords, cell_data):
    sen = cell_data.lower()
    for key in keywords:
        if re.search(key, sen):
            return True
    return False


def matcher_split(keywords, cell_data):
    clean_sentance = re.sub(r'[^\w]', ' ', cell_data.lower())
    word_list = clean_sentance.split()
    for key in keywords:
        if key in word_list:
            return True
    return False


class Tc_sorter:
    def __init__(self, test_case_list, output_name, last_week, difficult_list):
        print('Initiallizing...')
        self.test_case_list = str(test_case_list)
        self.output_name = str(output_name)
        self.last_week = str(last_week)
        self.sheet = (load_workbook(self.test_case_list)).active
        print('{} loaded successfully'.format(self.test_case_list))

        self.last_week_result = (
            load_workbook(self.last_week)).active
        print('{} loaded successfully'.format(self.last_week))

        self.difficult_list = str(difficult_list)
        self.dc_sheet = load_workbook(self.difficult_list).active
        print(print('{} loaded successfully'.format(self.difficult_list)))

        self.wb = Workbook()
        self.wb.active
        for name in sheet_names:
            self.wb.create_sheet(name, int((sheet_names).index(name)))
            self.wb[name].append(titles)
        for fail_name in fail_case_sheet:
            self.wb.create_sheet(fail_name, -1)
            self.wb[fail_name].append(fail_case_title)
        print('Output file initiallized')

    def difficult_cases(self):
        difficult_cases = []
        for row in self.dc_sheet.iter_rows(max_col=1, values_only=True):
            cells = []
            for cell in row:
                if cell is None:
                    cells.append('-')
                else:
                    cells.append(cell)
            if cells[-1] != '-':
                difficult_cases.append(cells[-1])
            elif cells[-1] == '-':
                break
        return difficult_cases[1:]

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def phone_type(self, cell_data):
        iphone = ['iphone', 'cp', 'wcp']
        android = ['android', 'waa', 'aa']
        phone_requirement = [0, 0]
        for cell in cell_data[5:7]:
            if matcher_split(iphone, cell):
                phone_requirement[0] = 1
            if matcher_slice(android, cell):
                phone_requirement[1] = 1
        if phone_requirement == [1, 0]:
            cell_data.append('iPhone')
        elif phone_requirement == [0, 1]:
            cell_data.append('Android')
        elif phone_requirement == [1, 1]:
            cell_data.append('Both')
        else:
            cell_data.append(' ')

    def sign_status(self, cell_data):
        sign_out = ['sign out', 'sign-out', 'signout',
                    'signed out', 'no google user is logged  in', 'No user is signed in']
        if matcher_slice(sign_out, cell_data[5]):
            cell_data.append('sign_out')
        else:
            cell_data.append('sign_in')

    def connection(self, cell_data):
        offline = ['offline']
        if matcher_split(offline, cell_data[5]):
            cell_data.append('Offline')
        else:
            cell_data.append('Online')

    def formatter(self, cell_data):
        for _ in range(4):
            cell_data.insert(1, '')

    def user(self, cell_data):
        guest = ['guest']
        others = ['secondary', 'user 1', 'user 2', 'user1', 'user2']
        primary = ['primary']
        if matcher_split(guest, cell_data[5]):
            cell_data.append('Guest')
        elif matcher_slice(others, cell_data[5]):
            cell_data.append('Others')
        elif matcher_split(guest, cell_data[6]) and (matcher_slice(others, cell_data[6]) or matcher_split(primary, cell_data[6])):
            cell_data.append('multiple')
        else:
            cell_data.append('Driver')

    def bench_only(self, cell_data):
        press_button = ['long press', 'short press', 'press "end" key']
        cluster = ['cluster', 'swc', 'ipc', 'clustor']
        speed_limit = ['speed limit']
        expection = ['short press Power key', 'Long press Power button',
                     'DLM', 'short press selection buttion on the rotary wheel']
        bench_only_case = False
        for cell in cell_data[5:8]:
            if (matcher_slice(press_button, cell) or matcher_slice(cluster, cell) or matcher_slice(speed_limit, cell)) and matcher_slice(expection, cell) != True:
                bench_only_case = True
        return bench_only_case

    def ac_only(self, cell_data):
        ac = ['a/c', 'temperature', 'climate',
              'defroster', 'hvac']
        ac_split = ['air', 'fan']
        ac_case = False
        for cell in cell_data[5:8]:
            if matcher_slice(ac, cell) or matcher_split(ac_split, cell):
                ac_case = True
        return ac_case

    def tc_location_dict(self):
        tc_location = load_workbook('TC_location.xlsx').active
        # stored the data in a dictionary {test_case: location}
        return {TCID: location for (TCID, location) in tc_location.iter_rows(
            max_col=2, values_only=True) if TCID is not None}

    def last_week_result_dict(self):
        last_week_dict = {}
        last_week = self.last_week_result
        for last_week_row in last_week.iter_rows(max_col=5, values_only=True):
            last_week_cell = self.cell_data(last_week_row)
            last_week_dict[last_week_cell[0]] = last_week_cell[1:]
        return last_week_dict

    def sorting(self):
        print('Opening a new sheet...')
        sheet = self.sheet
        print('Last week result loaded successfully')
        difficult_cases_list = self.difficult_cases()
        print('Difficult case list generated')
        location_dict = self.tc_location_dict()
        print('Test case location dictionary generated')
        print('Iterating through the test plan......')

        k = 1

        # Iterate through the unprocessd test cases
        # Only getting the first 5 values of each row (tc, precondition, test_steps, expected_result, test_objective}
        for row in sheet.iter_rows(max_col=5, values_only=True):
            print('Iterate case no. {}'.format(k))
            # turn the data into a list
            cell_data = self.cell_data(row)
            # adding 'pass/fail', 'Tester', 'Automation_comment', 'bug ID', 'Note' to the list
            self.formatter(cell_data)
            # determine the phone type
            self.phone_type(cell_data)
            # determine the user type
            self.user(cell_data)
            # determine online/offline
            self.connection(cell_data)
            # determine sign-in/sign-out
            self.sign_status(cell_data)
            k += 1

            print(cell_data)

            if cell_data[0] != 'none':
                cell_data.append(location_dict[cell_data[0]])
            elif cell_data[0] == 'none':
                break

            # for last_week_row in last_week.iter_rows(max_col=5, values_only=True):
            #     last_week_cell = self.cell_data(last_week_row)
            #     if last_week_cell[0] == cell_data[0]:
            #         cell_data.append(last_week_cell[1])
            #         cell_data.append(last_week_cell[2])
            #         cell_data.append(last_week_cell[3])
            #         cell_data.append(last_week_cell[4])
            last_week_dict = self.last_week_result_dict()

            for i in range(4):
                cell_data.append(last_week_dict[cell_data[0]][i])

            cell_data = cell_data[:5] + [cell_data[-1]] + cell_data[5:-1]

            # Distributing the test case to the desinated sheet
            if cell_data[0] in difficult_cases_list:
                self.wb['Difficult_cases'].append(cell_data)

            elif self.bench_only(cell_data):
                self.wb['Bench_only'].append(cell_data)

            elif self.ac_only(cell_data):
                self.wb['ac_only'].append(cell_data)

            else:
                i = 11
                if cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Online_In'].append(cell_data)
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Online_Out'].append(cell_data)
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_in':
                    self.wb['Driver_Offline_In'].append(cell_data)
                elif cell_data[i] == 'Driver' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_out':
                    self.wb['Driver_Offline_Out'].append(cell_data)

                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_in':
                    self.wb['Guest_Online_In'].append(cell_data)
                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Online' and cell_data[i+2] == 'sign_out':
                    self.wb['Guest_Online_Out'].append(cell_data)
                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_in':
                    self.wb['Guest_Offline_In'].append(cell_data)
                elif cell_data[i] == 'Guest' and cell_data[i+1] == 'Offline' and cell_data[i+2] == 'sign_out':
                    self.wb['Guest_Offline_Out'].append(cell_data)
                else:
                    self.wb['Other'].append(cell_data)

        print('Saving the file named {}'.format(self.output_name))
        self.wb.save(self.output_name)


testing = Tc_sorter('W51_test_plan.xlsx',
                    'output.xlsx', 'W51_result.xlsx', 'W51_difficult.xlsx')

testing.sorting()
