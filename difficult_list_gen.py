import re
from openpyxl import load_workbook
from openpyxl import Workbook

difficult_test = load_workbook('W51_MY22_Difficult_cases.xlsx')
sheet = difficult_test.active
difficult_list = []

for row in sheet.iter_rows(max_col=1, values_only=True):
    cells = []
    for cell in row:
        if cell is None:
            cells.append('-')
        else:
            cells.append(cell)
    if cells[-1] != '-':
        difficult_list.append(cells[-1])
    elif cells[-1] == '-':
        break

print(len(difficult_list))
