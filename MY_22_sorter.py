from openpyxl import load_workbook
from openpyxl import Workbook
import re

sheet_names = [
    'Difficult_cases', 'Bench_only', 'ac_only',
    'Driver_Online_In', 'Driver_Online_Out', 'Driver_Offline_In', 'Driver_Offline_Out',
    'Guest_Online_In', 'Guest_Online_Out', 'Guest_Offline_In', 'Guest_Offline_Out',
    'Other']

difficult_cases = ['TC_MFL_45104_Wireless_AndroidAuto_0058', 'TC_MFL_GAS_Google_Assistant_0636', 'TC_MFL_GAS_Google_Assistant_0652', 'TC_MFL_GAS_Google_Assistant_0774', 'TC_MFL_GAS_Google_Assistant_0692', 'TC_MFL_GAS_Google_Assistant_0538', 'TC_MFL_GAS_Google_Assistant_0623', 'TC_MFL_GAS_Google_Assistant_0628', 'TC_MFL_GAS_Google_Assistant_0767', 'TC_MFL_000000_GAS_Google_Assistant_0002', 'TC_MFL_GAS_Google_Assistant_0172', 'TC_MFL_GAS_Google_Assistant_0173', 'TC_MFL_GAS_Google_Assistant_0174', 'TC_MFL_GAS_Google_Assistant_0175', 'TC_MFL_GAS_Google_Assistant_0352', 'TC_MFL_GAS_Google_Assistant_0356', 'TC_MFL_GAS_Google_Assistant_0357', 'TC_MFL_GAS_Google_Assistant_0358', 'TC_MFL_GAS_Google_Assistant_0384', 'TC_MFL_GAS_Google_Assistant_0388', 'TC_MFL_GAS_Google_Assistant_0389', 'TC_MFL_GAS_Google_Assistant_0435', 'TC_MFL_GAS_Google_Assistant_0446', 'TC_MFL_GAS_Google_Assistant_0447', 'TC_MFL_GAS_Google_Assistant_0467', 'TC_MFL_GAS_Google_Assistant_0468', 'TC_MFL_GAS_Google_Assistant_0553', 'TC_MFL_GAS_Google_Assistant_0578', 'TC_MFL_GAS_Google_Assistant_0585', 'TC_MFL_GAS_Google_Assistant_0590', 'TC_MFL_GAS_Google_Assistant_0599', 'TC_MFL_GAS_Google_Assistant_0603', 'TC_MFL_GAS_Google_Assistant_0611', 'TC_MFL_GAS_Google_Assistant_0612', 'TC_MFL_GAS_Google_Assistant_0615', 'TC_MFL_GAS_Google_Assistant_0625', 'TC_MFL_GAS_Google_Assistant_0626', 'TC_MFL_GAS_Google_Assistant_0627', 'TC_MFL_GAS_Google_Assistant_0632', 'TC_MFL_GAS_Google_Assistant_0633', 'TC_MFL_GAS_Google_Assistant_0634', 'TC_MFL_GAS_Google_Assistant_0635', 'TC_MFL_GAS_Google_Assistant_0637', 'TC_MFL_GAS_Google_Assistant_0639', 'TC_MFL_GAS_Google_Assistant_0640', 'TC_MFL_GAS_Google_Assistant_0641', 'TC_MFL_GAS_Google_Assistant_0643', 'TC_MFL_GAS_Google_Assistant_0646', 'TC_MFL_GAS_Google_Assistant_0647', 'TC_MFL_GAS_Google_Assistant_0650', 'TC_MFL_GAS_Google_Assistant_0653', 'TC_MFL_GAS_Google_Assistant_0654', 'TC_MFL_GAS_Google_Assistant_0655',
                   'TC_MFL_GAS_Google_Assistant_0667', 'TC_MFL_GAS_Google_Assistant_0679', 'TC_MFL_GAS_Google_Assistant_0681', 'TC_MFL_GAS_Google_Assistant_0688', 'TC_MFL_GAS_Google_Assistant_0691', 'TC_MFL_GAS_Google_Assistant_0794', 'TC_Wireless_Android_Auto_193939_0011', 'TC_Android_Auto_0104_Wireless', 'TC_Android_Auto_0146_Wireless', 'TC_Android_Auto_0213_Wireless', 'TC_Android_Auto_0402_Wireless', 'TC_Android_Auto_Power_Off_0006', 'TC_Android_Auto_Wireless_210569_0003', 'TC_Android_Auto_connection_wireless_0001', 'TC_MEDIA_ANDROID_AUTO_0123_Wireless', 'TC_MFL_000000_GAS_Maps_0087', 'TC_MFL_000000_GAS_Maps_0101', 'TC_MFL_000000_GAS_Maps_0102', 'TC_MFL_000000_GAS_Maps_0106', 'TC_MFL_107206_Android_Auto_0070', 'TC_MFL_107206_Android_Auto_0071', 'TC_MFL_45104_Wireless_AndroidAuto_0038', 'TC_MFL_51771_BT_SR_Unavailable_0007', 'TC_MFL_51771_BT_SR_Unavailable_0008', 'TC_MFL_51771_BT_SR_Unavailable_0009', 'TC_MFL_51771_BT_SR_Unavailable_0016', 'TC_MFL_GAS_Google_Assistant_0201', 'TC_MFL_GAS_Google_Assistant_0203', 'TC_MFL_GAS_Google_Assistant_0204', 'TC_MFL_GAS_Google_Assistant_0205', 'TC_MFL_GAS_Google_Assistant_0206', 'TC_MFL_GAS_Google_Assistant_0353', 'TC_MFL_GAS_Google_Assistant_0385', 'TC_MFL_GAS_Google_Assistant_0436', 'TC_MFL_GAS_Google_Assistant_0593', 'TC_MFL_GAS_Google_Assistant_0629', 'TC_MFL_GAS_Google_Assistant_0695', 'TC_MFL_GAS_Google_Assistant_0696', 'TC_MFL_GAS_Google_Assistant_0697', 'TC_MFL_GAS_Google_Assistant_0729', 'TC_MFL_GAS_Maps_0160', 'TC_MFL_GAS_Maps_0161', 'TC_Maps_Card_0003', 'TC_PINUIN_IgnitionCycle_014', 'TC_PINUIN_IgnitionCycle_023', 'TC_PINUI_SwitchUser_022', 'TC_PINUI_SwitchUser_023', 'TC_PINUI_SwitchUser_024', 'TC_PINUI_SwitchUser_025', 'TC_PINUI_SwitchUser_026', 'TC_Projection_PTT_0008', 'TC_Wireless_AA_TermsofUse_13', 'TC_Wireless_Android_Auto_Interaction_003', 'TC_dialogue_0016', 'TC_dialogue_0020', 'TC_CLU_ANDROID_AUTO_0004_Wireless', 'TC_CLU_ANDROID_AUTO_0005_Wireless', 'TC_CLU_ANDROID_AUTO_0006_Wireless']

titles = ['Original GM TC ID', 'Pass/Fail', 'Tester', 'Automation Comment', 'Precondition',
          'Test steps', 'Expected', 'Testing Objective', 'Phone', 'User', 'Online/Offline', 'Sign Status',
          'W50_tester', 'W50_Comment', 'W50_result', 'W50_comment']


def matcher_slice(keywords, cell_data):
    sen = cell_data.lower()
    for key in keywords:
        if re.search(key, sen):
            return True
    return False


def matcher_split(keywords, cell_data):
    clean_sentance = re.sub(r'[^\w]', ' ', cell_data.lower())
    word_list = clean_sentance.split()
    for key in keywords:
        if key in word_list:
            return True
    return False


class Tc_sorter:
    def __init__(self, test_case_list, output_name, last_week):
        self.test_case_list = str(test_case_list)
        self.output_name = str(output_name)
        self.last_week = str(last_week)
        self.sheet = (load_workbook(self.test_case_list)).active
        self.last_week_result = (
            load_workbook(self.last_week)).active
        self.wb = Workbook()
        self.wb.active
        for name in sheet_names:
            self.wb.create_sheet(name, int((sheet_names).index(name)))
            self.wb[name].append(titles)

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def phone_type(self, cell_data):
        iphone = ['iphone', 'cp', 'wcp']
        android = ['android', 'waa', 'aa']
        phone_requirement = [0, 0]
        for cell in cell_data[4:6]:
            if matcher_split(iphone, cell):
                phone_requirement[0] = 1
            if matcher_slice(android, cell):
                phone_requirement[1] = 1
        if phone_requirement == [1, 0]:
            cell_data.append('iPhone')
        elif phone_requirement == [0, 1]:
            cell_data.append('Android')
        elif phone_requirement == [1, 1]:
            cell_data.append('Both')
        else:
            cell_data.append(' ')

    def sign_status(self, cell_data):
        sign_out = ['sign out', 'sign-out', 'signout',
                    'signed out', 'no google user is logged  in', 'No user is signed in']
        if matcher_slice(sign_out, cell_data[4]):
            cell_data.append('sign_out')
        else:
            cell_data.append('sign_in')

    def connection(self, cell_data):
        offline = ['offline']
        if matcher_split(offline, cell_data[4]):
            cell_data.append('Offline')
        else:
            cell_data.append('Online')

    def formatter(self, cell_data):
        for _ in range(3):
            cell_data.insert(1, '')

    def user(self, cell_data):
        guest = ['guest']
        others = ['secondary', 'user 1', 'user 2', 'user1', 'user2']
        primary = ['primary']
        if matcher_split(guest, cell_data[4]):
            cell_data.append('Guest')
        elif matcher_slice(others, cell_data[4]):
            cell_data.append('Others')
        elif matcher_split(guest, cell_data[5]) and (matcher_slice(others, cell_data[5]) or matcher_split(primary, cell_data[5])):
            cell_data.append('multiple')
        else:
            cell_data.append('Driver')

    def bench_only(self, cell_data):
        press_button = ['long press', 'short press', 'press "end" key']
        cluster = ['cluster', 'swc']
        speed_limit = ['speed limit']
        expection = ['short press Power key', 'Long press Power button', 'DLM']
        bench_only_case = False
        for cell in cell_data[4:7]:
            if (matcher_slice(press_button, cell) or matcher_slice(cluster, cell) or matcher_slice(speed_limit, cell)) and matcher_slice(expection, cell) != True:
                bench_only_case = True
        return bench_only_case

    def ac_only(self, cell_data):
        ac = ['a/c', 'temperature', 'climate',
              'defroster', 'hvac']
        ac_split = ['air', 'fan']
        ac_case = False
        for cell in cell_data[4:7]:
            if matcher_slice(ac, cell) or matcher_split(ac_split, cell):
                ac_case = True
        return ac_case

    def sorting(self):
        sheet = self.sheet
        last_week = self.last_week_result
        for row in sheet.iter_rows(max_col=5, values_only=True):
            cell_data = self.cell_data(row)
            self.formatter(cell_data)
            self.phone_type(cell_data)
            self.user(cell_data)
            self.connection(cell_data)
            self.sign_status(cell_data)

            for last_week_row in last_week.iter_rows(max_col=5, values_only=True):
                last_week_cell = self.cell_data(last_week_row)
                if last_week_cell[0] == cell_data[0]:
                    cell_data.append(last_week_cell[2])
                    cell_data.append(last_week_cell[3])
                    cell_data.append(last_week_cell[1])
                    cell_data.append(last_week_cell[4])

            # the final format will be like this:
            # ['ID', 'Pass/Fail', 'tester', 'comment',
            #  'precondition', 'test_steps', 'expected_result',
            #  'phone_type', 'user', 'connection', 'sign_status',
            #  'name of tester', 'last_week_result']

            if cell_data[0] in difficult_cases:
                self.wb['Difficult_cases'].append(cell_data)

            elif self.bench_only(cell_data):
                self.wb['Bench_only'].append(cell_data)

            elif self.ac_only(cell_data):
                self.wb['ac_only'].append(cell_data)

            else:

                if cell_data[9] == 'Driver' and cell_data[10] == 'Online' and cell_data[11] == 'sign_in':
                    self.wb['Driver_Online_In'].append(cell_data)
                elif cell_data[9] == 'Driver' and cell_data[10] == 'Online' and cell_data[11] == 'sign_out':
                    self.wb['Driver_Online_Out'].append(cell_data)
                elif cell_data[9] == 'Driver' and cell_data[10] == 'Offline' and cell_data[11] == 'sign_in':
                    self.wb['Driver_Offline_In'].append(cell_data)
                elif cell_data[9] == 'Driver' and cell_data[10] == 'Offline' and cell_data[11] == 'sign_out':
                    self.wb['Driver_Offline_Out'].append(cell_data)

                elif cell_data[9] == 'Guest' and cell_data[10] == 'Online' and cell_data[11] == 'sign_in':
                    self.wb['Guest_Online_In'].append(cell_data)
                elif cell_data[9] == 'Guest' and cell_data[10] == 'Online' and cell_data[11] == 'sign_out':
                    self.wb['Guest_Online_Out'].append(cell_data)
                elif cell_data[9] == 'Guest' and cell_data[10] == 'Offline' and cell_data[11] == 'sign_in':
                    self.wb['Guest_Offline_In'].append(cell_data)
                elif cell_data[9] == 'Guest' and cell_data[10] == 'Offline' and cell_data[11] == 'sign_out':
                    self.wb['Guest_Offline_Out'].append(cell_data)
                else:
                    self.wb['Other'].append(cell_data)

        self.wb.save(self.output_name)


testing = Tc_sorter('W51.xlsx',
                    'MY22_W51.xlsx', 'Result_P&F_W50.xlsx')

testing.sorting()
