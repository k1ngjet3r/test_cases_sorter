from openpyxl import load_workbook, Workbook

# create an output worksheet
result = Workbook()
result.create_sheet('match result')
result['match result'].append(['Case ID', 'MY22', 'MY23', 'Automated'])

result.create_sheet('Summary')
result['Summary'].append(['MY22 & MY23 Intersect', 'MY22 Automated', 'MY23 Automated', 'Both Automated'])

# for storing all cases
all_case_list = []

# loading my_22 case list and generate a list of case ID
my_22 = load_workbook('MY_22.xlsx').active
my_22_list =  []
for i in my_22.iter_rows(max_col=1, values_only=True):
    if not i[0] == None:
        my_22_list.append(i[0].lower())
        if i[0].lower() not in all_case_list:
            all_case_list.append(i[0].lower())

# loading my_23 case list and generate a list of case ID
my_23 = load_workbook('MY_23.xlsx').active
my_23_list = []
for j in my_23.iter_rows(max_col=1, values_only=True):
    if not j[0] == None:
        my_23_list.append(j[0].lower())
        if j[0].lower() not in all_case_list:
            all_case_list.append(j[0].lower())

# loading automatic cases
auto_cases = load_workbook('all_auto_cases.xlsx').active
auto_cases_list = []
for k in auto_cases.iter_rows(max_col=1, values_only=True):
    if not k[0] == None:
        auto_cases_list.append(k[0].lower())

num_case = len(all_case_list)
current = 0

my_22_and_23 = 0
my_22_automated = 0
my_23_automated = 0
both_automated = 0

for case in all_case_list:
    current += 1
    print('Pregress: {}/{}'.format(current, num_case))
    case_rlt = [case]
    if case in my_22_list:
        case_rlt.append('Y')
    else:
        case_rlt.append('')

    if case in my_23_list:
        case_rlt.append('Y')
    else:
        case_rlt.append('')

    if case in auto_cases_list:
        case_rlt.append('Y')
    else:
        case_rlt.append('')

    if case_rlt[1] == 'Y' and case_rlt[2] == 'Y':
        my_22_and_23 += 1
    if case_rlt[1] == 'Y' and case_rlt[3] == 'Y':
        my_22_automated += 1
    if case_rlt[2] == 'Y' and case_rlt[3] == 'Y':
        my_23_automated += 1
    if case_rlt[1] == 'Y' and case_rlt[2] == 'Y' and case_rlt[3] == 'Y':
        both_automated += 1

    result['match result'].append(case_rlt)
    result.save('case_comparison.xlsx')

summary_result = [my_22_and_23, my_22_automated, my_23_automated, both_automated]
result['Summary'].append(summary_result)

