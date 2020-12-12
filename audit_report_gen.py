from openpyxl import load_workbook
from openpyxl import Workbook
import re


class Tc:
    def __init__(self, input_name):
        self.input_name = str(input_name)
        self.isheet = (load_workbook(self.input_name)).active
        self.wb = Workbook()
        self.osheet = self.wb.active

    def cell_data(self, row):
        cells = []
        for cell in row:
            if cell is None:
                cells.append('none')
            else:
                cells.append(cell)
        return cells

    def matching(self, wb2, output_name):
        wb1 = self
        isheet1 = wb1.isheet
        isheet2 = wb2.isheet
        for row1 in isheet1.iter_rows(max_col=2, values_only=True):
            cell_data = self.cell_data(row1)
            for row2 in isheet2.iter_rows(max_col=2, values_only=True):
                cell_temp = self.cell_data(row2)
                if cell_data[0].lower() == cell_temp[0].lower():
                    cell_data.append(cell_temp[1])
                    wb1.osheet.append(cell_data)
        wb1.wb.save(str(output_name))


test1 = Tc('test1.xlsx')
test2 = Tc('test2.xlsx')
test1.matching(test2, 'test.xlsx')
